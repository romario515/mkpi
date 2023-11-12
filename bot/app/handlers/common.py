from aiogram import Dispatcher, types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text, IDFilter
from aiogram.dispatcher.filters.state import State, StatesGroup
from emoji import emojize
import openpyxl
import time
import datetime
import pytz
import schedule
SECRET = None
database = None


class Common(StatesGroup):
    waiting_for_passphrase = State()


def generate_keyboard(chat_id=None) -> types.ReplyKeyboardMarkup:
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    admins = database.get_admin_list()
    if chat_id and str(chat_id) in admins:
        keyboard.add("Статус точек")
        keyboard.add("Статистика за день", "Архив")
    else:
        keyboard.add("Регистрация")
    return keyboard


async def cmd_start(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer(
        "Выберите действие",
        reply_markup=generate_keyboard(message.chat.id)
        
    )


async def cmd_stats(message: types.Message, state: FSMContext):
    await state.finish()
    stat_msg = 'Статус точек:\n'
    left_pos = database.get_work_status_left()
    #print(left_pos[0])
    stat_msg += f"Точка {left_pos[0]}:   {left_pos[1]} \n"
    right_pos = database.get_work_status_right()
    stat_msg += f"Точка {right_pos[0]}:   {right_pos[1]} \n"
    await message.answer(stat_msg, reply_markup=types.ReplyKeyboardRemove())
    await message.answer(
        "Выберите действие",
        reply_markup=generate_keyboard(message.chat.id)
    )

async def cmd_stat_dot(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer(
        "Статистика времени за день:\n",
        reply_markup=types.ReplyKeyboardRemove()
    )
    prostoy=database.get_day_stat('prostoy','left')
    print (prostoy)
    pros_ini=time.strftime('%H:%M:%S', time.gmtime(int(prostoy)))
    msg='На простой затрачено: '+ pros_ini +'\n'
    ving=database.get_day_stat('vinuzhdennaya','left')
    svarkag = database.get_day_stat('svarka','left')
    vin=time.strftime('%H:%M:%S', time.gmtime(int(ving)))
    svarka=time.strftime('%H:%M:%S', time.gmtime(int(svarkag)))
    msg+='На вынужденную работу затрачено: '+ vin +'\n'
    msg+='На сварку затрачено: '+ svarka +'\n'
    await message.answer(
        msg,
        reply_markup=generate_keyboard(message.chat.id)
    )


async def cmd_cancel(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer("Действие отменено", reply_markup=types.ReplyKeyboardRemove())
    await message.answer(
        "Выберите действие",
        reply_markup=generate_keyboard(message.chat.id))

    left = database.sort_history_left()
    right = database.sort_history_right()
     # Создание нового Excel документа
    wb = openpyxl.Workbook()
    sheet = wb.active
    # Заполнение данных
    sheet.cell(row=1, column=1, value='Время')
    sheet.cell(row=1, column=2, value='Причина')
    sheet.cell(row=1, column=4, value='Время')
    sheet.cell(row=1, column=5, value='Причина')
    for i, (val1, val2) in enumerate(left):
        sheet.cell(row=i+2, column=1, value=val2)
        sheet.cell(row=i+2, column=2, value=val1)
    for i, (val1, val2) in enumerate(right):
        sheet.cell(row=i+2, column=4, value=val2)
        sheet.cell(row=i+2, column=5, value=val1)
        # Сохранение файла
    wb.save('output.xlsx')
    

async def job(message: types.Message, state: FSMContext):
    moscow_time = datetime.now(pytz.timezone('Europe/Moscow'))
    if moscow_time.strftime("%H:%M") == "20:00":
        left = database.get_sorted_history_left()
        right = database.get_sorted_history_right()

        # Создание нового Excel документа
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Заполнение данных
        sheet.cell(row=1, column=1, value='Время')
        sheet.cell(row=1, column=2, value='Причина')
        sheet.cell(row=1, column=4, value='Время')
        sheet.cell(row=1, column=5, value='Причина')
        for i, (val1, val2) in enumerate(left):
            sheet.cell(row=i+2, column=1, value=val2)
            sheet.cell(row=i+2, column=2, value=val1)
        for i, (val1, val2) in enumerate(right):
            sheet.cell(row=i+2, column=4, value=val2)
            sheet.cell(row=i+2, column=5, value=val1)
        # Сохранение файла
        wb.save('output.xlsx')
    # Настройка расписания для выполнения задачи каждый день в 20:0
    schedule.every().day.at("20:00").do(job)
    while True:
        schedule.run_pending()
        time.sleep(60)  # Проверка каждую минуту

    


async def cmd_evolve(message: types.Message, state: FSMContext):
    await message.answer("Введите код аутентификации",
                         reply_markup=types.ReplyKeyboardRemove())
    await state.set_state(Common.waiting_for_passphrase.state)


async def evolve_pass(message: types.Message, state: FSMContext):
    if message.text.lower() == SECRET:
        database.become_admin(message.chat.id)
        await message.answer("Регистрация пройдена успешно!\n/start для возврата в меню",
                             reply_markup=types.ReplyKeyboardRemove())
        await state.finish()
    else:
        await message.answer("Не верный код, повторите попытку",
                             reply_markup=types.ReplyKeyboardRemove())


def register_handlers_common(dp: Dispatcher, secret: str, db):
    global SECRET, database
    SECRET = secret
    database = db
    dp.register_message_handler(cmd_start, commands="start", state="*")
    dp.register_message_handler(cmd_start, Text(equals="Возврат в меню", ignore_case=True), state="*")
    dp.register_message_handler(cmd_cancel, commands="cancel", state="*")
    dp.register_message_handler(cmd_cancel, Text(equals="Отмена", ignore_case=True), state="*")
    dp.register_message_handler(cmd_stats, commands="statistic", state="*")
    dp.register_message_handler(cmd_stats, Text(equals="Статус точек", ignore_case=True),
                                state="*")
    dp.register_message_handler(cmd_stat_dot, commands="stat_dot", state="*")
    dp.register_message_handler(cmd_stat_dot, Text(equals="Статистика за день", ignore_case=True), state="*")

    dp.register_message_handler(cmd_evolve, commands="evolve", state="*")
    dp.register_message_handler(cmd_evolve, Text(equals="Регистрация", ignore_case=True), state="*")
    dp.register_message_handler(evolve_pass, state=Common.waiting_for_passphrase)
    dp.register_message_handler(job, commands="job", state="*")
    
    

